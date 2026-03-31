"""Jaarcijfers router - generates annual financial report data."""

import csv
import io
import zipfile
from datetime import date
from urllib.parse import urlparse, unquote

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from firebase_admin import firestore, storage
from google.cloud.firestore_v1 import FieldFilter
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers

from app.auth import get_current_user
from app.services.pdf_generator import generate_invoice_pdf
from app.config import FIREBASE_STORAGE_BUCKET

router = APIRouter()


def get_db():
    return firestore.client()


def get_year(date_str: str) -> int:
    try:
        return int(date_str[:4])
    except (ValueError, IndexError):
        return 0


# === Core computation ===

def _compute_jaarcijfers(
    jaar: int,
    all_invoice_data: list[dict],
    all_expense_data: list[dict],
    bank_accounts: dict | None = None,
    prev_year_eind: dict | None = None,
) -> dict:
    """Compute full jaarcijfers for a single year. Pure computation, no DB calls.
    
    prev_year_eind: if provided, overrides the computed begin-of-year values
    with the previous year's eind values from accountant data.
    Expected keys: mva, debiteuren, liquide_middelen, crediteuren, btw_schuld, eigen_vermogen
    """

    # === 1. NETTO-OMZET ===
    omzet = 0.0
    omzet_btw = 0.0
    omzet_per_klant = defaultdict(float)
    for inv in all_invoice_data:
        if get_year(inv.get("factuurdatum", "")) == jaar and inv.get("status") in ("verzonden", "betaald"):
            omzet += inv.get("subtotaal", 0)
            omzet_btw += inv.get("btw_totaal", 0)
            klant = inv.get("klant_naam", "Onbekend") or "Onbekend"
            omzet_per_klant[klant] += inv.get("subtotaal", 0)

    # === 2. KOSTEN & AFSCHRIJVINGEN ===
    kosten_direct = 0.0
    afschrijvingen = 0.0
    kosten_per_categorie = defaultdict(float)
    afschrijving_items = []

    for exp in all_expense_data:
        exp_year = get_year(exp.get("datum", ""))
        is_afschrijving = exp.get("afschrijving", False)

        if is_afschrijving:
            jaren = exp.get("afschrijving_jaren") or 1
            restwaarde = exp.get("afschrijving_restwaarde") or 0
            subtotaal = exp.get("subtotaal", 0)
            jaarlijks = (subtotaal - restwaarde) / jaren

            if exp_year <= jaar < exp_year + jaren:
                afschrijvingen += jaarlijks
                afschrijving_items.append({
                    "id": exp.get("id", ""),
                    "leverancier": exp.get("leverancier", ""),
                    "beschrijving": exp.get("beschrijving", ""),
                    "datum": exp.get("datum", ""),
                    "categorie": exp.get("categorie", ""),
                    "aanschafwaarde": subtotaal,
                    "restwaarde": restwaarde,
                    "jaren": jaren,
                    "jaarlijkse_afschrijving": round(jaarlijks, 2),
                    "jaar_van_aanschaf": exp_year,
                    "boekwaarde_begin": round(max(0, subtotaal - jaarlijks * (jaar - exp_year)), 2),
                    "boekwaarde_eind": round(max(0, subtotaal - jaarlijks * (jaar - exp_year + 1)), 2),
                })
        else:
            if exp_year == jaar:
                subtotaal = exp.get("subtotaal", 0)
                kosten_direct += subtotaal
                cat = exp.get("categorie", "Overig") or "Overig"
                kosten_per_categorie[cat] += subtotaal

    # === 3. MVA (Materiële Vaste Activa) ===
    mva_items = []
    mva_boekwaarde_begin = 0.0
    mva_boekwaarde_eind = 0.0
    mva_aanschaf_dit_jaar = 0.0

    for exp in all_expense_data:
        if not exp.get("afschrijving"):
            continue
        exp_year = get_year(exp.get("datum", ""))
        jaren = exp.get("afschrijving_jaren") or 1
        restwaarde = exp.get("afschrijving_restwaarde") or 0
        subtotaal = exp.get("subtotaal", 0)
        jaarlijks = (subtotaal - restwaarde) / jaren

        years_elapsed_begin = jaar - exp_year
        years_elapsed_end = jaar - exp_year + 1

        bw_begin = 0.0
        bw_eind = 0.0

        if 0 <= years_elapsed_begin <= jaren:
            bw_begin = max(0, subtotaal - jaarlijks * years_elapsed_begin)
        if 0 < years_elapsed_end <= jaren:
            bw_eind = max(0, subtotaal - jaarlijks * years_elapsed_end)
        if exp_year == jaar:
            bw_begin = 0
            mva_aanschaf_dit_jaar += subtotaal

        if bw_begin > 0 or bw_eind > 0:
            mva_items.append({
                "id": exp.get("id", ""),
                "leverancier": exp.get("leverancier", ""),
                "beschrijving": exp.get("beschrijving", ""),
                "datum": exp.get("datum", ""),
                "categorie": exp.get("categorie", ""),
                "aanschafwaarde": subtotaal,
                "restwaarde": restwaarde,
                "jaren": jaren,
                "jaarlijkse_afschrijving": round(jaarlijks, 2),
                "boekwaarde_begin": round(bw_begin, 2),
                "boekwaarde_eind": round(bw_eind, 2),
                "afschrijving_dit_jaar": round(jaarlijks if exp_year <= jaar < exp_year + jaren else 0, 2),
            })
            mva_boekwaarde_begin += bw_begin
            mva_boekwaarde_eind += bw_eind

    # === 4. DEBITEUREN ===
    # All invoices marked "betaald" are treated as paid on invoice date.
    # Only "verzonden" invoices count as outstanding.
    debiteuren_begin = 0.0
    debiteuren_eind = 0.0

    for inv in all_invoice_data:
        datum = inv.get("factuurdatum", "")
        inv_year = get_year(datum)
        status = inv.get("status", "")
        totaal = inv.get("totaal", 0)

        # Eind: facturen t/m dit jaar die nog niet betaald zijn
        if inv_year <= jaar and status == "verzonden":
            debiteuren_eind += totaal

        # Begin: facturen van vóór dit jaar die nog niet betaald zijn
        if inv_year < jaar and status == "verzonden":
            debiteuren_begin += totaal

    # === 5. CREDITEUREN ===
    calc_crediteuren_begin = 0.0
    calc_crediteuren_eind = 0.0
    PAID_STATUSES = {"betaald", "verwerkt"}

    for exp in all_expense_data:
        if exp.get("afschrijving"):
            continue
        datum = exp.get("datum", "")
        exp_year = get_year(datum)
        status = exp.get("status", "")
        totaal = exp.get("totaal", 0)

        if exp_year == jaar and status not in PAID_STATUSES:
            calc_crediteuren_eind += totaal
        if exp_year < jaar and status not in PAID_STATUSES:
            calc_crediteuren_begin += totaal

    crediteuren_begin = calc_crediteuren_begin
    crediteuren_eind = calc_crediteuren_eind

    # === 6. BTW SCHULD ===
    btw_schuld_begin = 0.0
    btw_schuld_eind = 0.0

    for inv in all_invoice_data:
        datum = inv.get("factuurdatum", "")
        if get_year(datum) == jaar and inv.get("status") in ("verzonden", "betaald"):
            try:
                month = int(datum[5:7])
                if month >= 10:
                    btw_schuld_eind += inv.get("btw_totaal", 0)
            except (ValueError, IndexError):
                pass

    for exp in all_expense_data:
        datum = exp.get("datum", "")
        if get_year(datum) == jaar:
            try:
                month = int(datum[5:7])
                if month >= 10:
                    btw_schuld_eind -= exp.get("btw", 0)
            except (ValueError, IndexError):
                pass

    for inv in all_invoice_data:
        datum = inv.get("factuurdatum", "")
        if get_year(datum) == jaar - 1 and inv.get("status") in ("verzonden", "betaald"):
            try:
                month = int(datum[5:7])
                if month >= 10:
                    btw_schuld_begin += inv.get("btw_totaal", 0)
            except (ValueError, IndexError):
                pass

    for exp in all_expense_data:
        datum = exp.get("datum", "")
        if get_year(datum) == jaar - 1:
            try:
                month = int(datum[5:7])
                if month >= 10:
                    btw_schuld_begin -= exp.get("btw", 0)
            except (ValueError, IndexError):
                pass

    # BTW: use computed values directly

    # MVA: use computed values directly
    final_mva_begin = mva_boekwaarde_begin
    final_mva_eind = mva_boekwaarde_eind

    # === 7. LIQUIDE MIDDELEN (from bank CSV data) ===
    liquide_middelen_begin = None
    liquide_middelen_eind = None
    has_bank_data = bool(bank_accounts)

    if bank_accounts:
        lm_begin = _compute_liquide_middelen(bank_accounts, f"{jaar - 1}-12-31")
        lm_eind = _compute_liquide_middelen(bank_accounts, f"{jaar}-12-31")
        if lm_begin is not None:
            liquide_middelen_begin = lm_begin
        if lm_eind is not None:
            liquide_middelen_eind = lm_eind

    # === 7b. APPLY PREVIOUS YEAR OVERRIDES for begin values ===
    if prev_year_eind:
        if "mva" in prev_year_eind:
            final_mva_begin = prev_year_eind["mva"]
        if "debiteuren" in prev_year_eind:
            debiteuren_begin = prev_year_eind["debiteuren"]
        if "liquide_middelen" in prev_year_eind:
            liquide_middelen_begin = prev_year_eind["liquide_middelen"]
        if "crediteuren" in prev_year_eind:
            crediteuren_begin = prev_year_eind["crediteuren"]
        if "btw_schuld" in prev_year_eind:
            btw_schuld_begin = prev_year_eind["btw_schuld"]

    # === 8. BALANS ===
    activa_begin = round(
        final_mva_begin + debiteuren_begin + (liquide_middelen_begin or 0), 2
    )
    activa_eind = round(
        final_mva_eind + debiteuren_eind + (liquide_middelen_eind or 0), 2
    )

    passiva_crediteuren_begin = round(crediteuren_begin, 2)
    passiva_crediteuren_eind = round(crediteuren_eind, 2)
    passiva_btw_begin = round(max(0, btw_schuld_begin), 2)
    passiva_btw_eind = round(max(0, btw_schuld_eind), 2)

    passiva_kort_begin = passiva_crediteuren_begin + passiva_btw_begin
    passiva_kort_eind = passiva_crediteuren_eind + passiva_btw_eind

    eigen_vermogen_begin = round(activa_begin - passiva_kort_begin, 2)
    eigen_vermogen_eind = round(activa_eind - passiva_kort_eind, 2)

    # === 8. WINSTBEREKENING ===
    totaal_kosten = kosten_direct + afschrijvingen
    winst = omzet - totaal_kosten

    # === 9. MKB-WINSTVRIJSTELLING ===
    mkb_percentages = {2022: 0.14, 2023: 0.14}
    mkb_pct = mkb_percentages.get(jaar, 0.1331)
    mkb_vrijstelling = round(winst * mkb_pct, 2)
    belastbare_winst = round(winst - mkb_vrijstelling, 2)

    return {
        "jaar": jaar,
        "winst_verlies": {
            "omzet": round(omzet, 2),
            "omzet_btw": round(omzet_btw, 2),
            "kosten_direct": round(kosten_direct, 2),
            "afschrijvingen": round(afschrijvingen, 2),
            "totaal_kosten": round(totaal_kosten, 2),
            "winst": round(winst, 2),
            "mkb_vrijstelling": mkb_vrijstelling,
            "belastbare_winst": belastbare_winst,
            "mkb_percentage": mkb_pct,
            "omzet_per_klant": sorted(
                [{"naam": k, "bedrag": round(v, 2)} for k, v in omzet_per_klant.items()],
                key=lambda x: x["bedrag"], reverse=True,
            ),
            "kosten_per_categorie": sorted(
                [{"naam": k, "bedrag": round(v, 2)} for k, v in kosten_per_categorie.items()],
                key=lambda x: x["bedrag"], reverse=True,
            ),
        },
        "balans": {
            "activa": {
                "mva": {"begin": round(final_mva_begin, 2), "eind": round(final_mva_eind, 2)},
                "debiteuren": {"begin": round(debiteuren_begin, 2), "eind": round(debiteuren_eind, 2)},
                "liquide_middelen": {
                    "begin": liquide_middelen_begin,
                    "eind": liquide_middelen_eind,
                },
                "totaal": {"begin": activa_begin, "eind": activa_eind},
            },
            "passiva": {
                "eigen_vermogen": {"begin": eigen_vermogen_begin, "eind": eigen_vermogen_eind},
                "crediteuren": {"begin": passiva_crediteuren_begin, "eind": passiva_crediteuren_eind},
                "btw_schuld": {"begin": passiva_btw_begin, "eind": passiva_btw_eind},
                "kortlopend_totaal": {"begin": passiva_kort_begin, "eind": passiva_kort_eind},
                "totaal": {"begin": round(eigen_vermogen_begin + passiva_kort_begin, 2), "eind": round(eigen_vermogen_eind + passiva_kort_eind, 2)},
            },
        },
        "mva": {
            "items": sorted(mva_items, key=lambda x: x.get("datum", "")),
            "totaal_boekwaarde_begin": round(mva_boekwaarde_begin, 2),
            "totaal_boekwaarde_eind": round(mva_boekwaarde_eind, 2),
            "totaal_afschrijving": round(afschrijvingen, 2),
            "totaal_aanschaf_dit_jaar": round(mva_aanschaf_dit_jaar, 2),
        },
        "bron": "berekend",
    }


def _load_all_data(db, uid):
    """Load all invoices, expenses, and determine available years."""
    invoices = list(
        db.collection("invoices")
        .where(filter=FieldFilter("user_id", "==", uid))
        .stream()
    )
    all_invoice_data = [{"id": doc.id, **doc.to_dict()} for doc in invoices]

    expenses = list(
        db.collection("expenses")
        .where(filter=FieldFilter("user_id", "==", uid))
        .stream()
    )
    all_expense_data = [{"id": doc.id, **doc.to_dict()} for doc in expenses]

    all_years = set()
    for inv in all_invoice_data:
        y = get_year(inv.get("factuurdatum", ""))
        if y:
            all_years.add(y)
    for exp in all_expense_data:
        y = get_year(exp.get("datum", ""))
        if y:
            all_years.add(y)
            if exp.get("afschrijving"):
                jaren = exp.get("afschrijving_jaren") or 1
                for offset in range(jaren):
                    all_years.add(y + offset)

    return all_invoice_data, all_expense_data, sorted(all_years, reverse=True)


# === Bank CSV Parsing ===

def _detect_and_parse_csv(content: str) -> tuple[str, str, list[dict]]:
    """
    Detect CSV format and parse transactions.
    Returns (account_name, account_number, transactions).

    Supports two ING formats:
    1. Betaalrekening: date=YYYYMMDD, amount col="Bedrag (EUR)", has "Code"/"Tag"
    2. Spaarrekening: date=YYYY-MM-DD, amount col="Bedrag", has "Rekening naam"/"Valuta"
    """
    reader = csv.DictReader(io.StringIO(content), delimiter=";")
    fields = reader.fieldnames or []
    # Strip quotes from fieldnames
    fields = [f.strip('"') for f in fields]
    reader.fieldnames = fields

    is_spaar = "Rekening naam" in fields
    amount_col = "Bedrag" if is_spaar else "Bedrag (EUR)"
    account_number = ""
    account_name = ""

    transactions = []
    for row in reader:
        raw_datum = row.get("Datum", "").strip('"')
        if is_spaar:
            # Format: YYYY-MM-DD → keep as-is
            datum = raw_datum
        else:
            # Format: YYYYMMDD → convert to YYYY-MM-DD
            if len(raw_datum) == 8:
                datum = f"{raw_datum[:4]}-{raw_datum[4:6]}-{raw_datum[6:8]}"
            else:
                datum = raw_datum

        af_bij = row.get("Af Bij", "").strip('"')
        bedrag_str = row.get(amount_col, "0").strip('"').replace(".", "").replace(",", ".")
        try:
            bedrag = float(bedrag_str)
        except ValueError:
            bedrag = 0.0
        if af_bij == "Af":
            bedrag = -bedrag

        saldo_str = row.get("Saldo na mutatie", "0").strip('"').replace(".", "").replace(",", ".")
        try:
            saldo = float(saldo_str)
        except ValueError:
            saldo = 0.0

        if not account_number:
            account_number = row.get("Rekening", "").strip('"')
        if is_spaar and not account_name:
            account_name = row.get("Rekening naam", "").strip('"')

        omschrijving = row.get("Naam / Omschrijving", "") or row.get("Omschrijving", "")
        omschrijving = omschrijving.strip('"')

        tegenrekening = row.get("Tegenrekening", "").strip('"')
        mededelingen = row.get("Mededelingen", "").strip('"')

        transactions.append({
            "datum": datum,
            "omschrijving": omschrijving,
            "bedrag": round(bedrag, 2),
            "saldo_na_mutatie": round(saldo, 2),
            "af_bij": af_bij,
            "mutatiesoort": row.get("Mutatiesoort", "").strip('"'),
            "tegenrekening": tegenrekening,
            "mededelingen": mededelingen,
        })

    if not account_name:
        account_name = f"Betaalrekening {account_number}"

    return account_name, account_number, transactions


def _get_saldo_at_date(transactions: list[dict], target_date: str) -> float | None:
    """
    Get account balance at a specific date (YYYY-MM-DD).
    Finds the last transaction on or before target_date.
    """
    best = None
    for tx in transactions:
        if tx["datum"] <= target_date:
            if best is None or tx["datum"] > best["datum"]:
                best = tx
            elif tx["datum"] == best["datum"]:
                best = tx  # last one in list wins if same date
    return best["saldo_na_mutatie"] if best else None


def _load_bank_data(db, uid) -> dict:
    """
    Load bank accounts and compute saldo per date from stored transactions.
    Returns: {account_number: {"name": ..., "transactions": [...], "min_date": ..., "max_date": ...}}
    """
    accounts = {}
    docs = list(
        db.collection("bank_accounts")
        .where(filter=FieldFilter("user_id", "==", uid))
        .stream()
    )
    for doc in docs:
        d = doc.to_dict()
        acc_nr = d.get("account_number", "")
        accounts[acc_nr] = {
            "name": d.get("account_name", acc_nr),
            "min_date": d.get("min_date", ""),
            "max_date": d.get("max_date", ""),
        }

    # Load transactions per account
    for acc_nr in accounts:
        tx_docs = list(
            db.collection("bank_transactions")
            .where(filter=FieldFilter("user_id", "==", uid))
            .where(filter=FieldFilter("account_number", "==", acc_nr))
            .stream()
        )
        txs = [doc.to_dict() for doc in tx_docs]
        # Sort by date
        txs.sort(key=lambda x: x.get("datum", ""))
        accounts[acc_nr]["transactions"] = txs

    return accounts


def _compute_liquide_middelen(bank_accounts: dict, target_date: str) -> float:
    """Sum saldo across all bank accounts at a given date."""
    total = 0.0
    for acc_nr, acc in bank_accounts.items():
        txs = acc.get("transactions", [])
        saldo = _get_saldo_at_date(txs, target_date)
        if saldo is not None:
            total += saldo
    return round(total, 2)


# === Override (accountant) data ===

def _load_overrides(db, uid) -> dict[int, dict]:
    """Load accountant override data per year from Firestore."""
    docs = list(
        db.collection("jaarcijfers_overrides")
        .where(filter=FieldFilter("user_id", "==", uid))
        .stream()
    )
    overrides = {}
    for doc in docs:
        d = doc.to_dict()
        jaar = d.get("jaar")
        if jaar:
            overrides[jaar] = d
    return overrides


def _override_to_jaarcijfers(override: dict) -> dict:
    """Convert a stored override to the same format as _compute_jaarcijfers output."""
    jaar = override["jaar"]
    wv = override.get("winst_verlies", {})
    bal = override.get("balans", {})
    mva_data = override.get("mva", {})

    return {
        "jaar": jaar,
        "winst_verlies": {
            "omzet": wv.get("omzet", 0),
            "omzet_btw": wv.get("omzet_btw", 0),
            "kosten_direct": wv.get("kosten_direct", 0),
            "afschrijvingen": wv.get("afschrijvingen", 0),
            "totaal_kosten": wv.get("totaal_kosten", 0),
            "winst": wv.get("winst", 0),
            "mkb_vrijstelling": wv.get("mkb_vrijstelling", 0),
            "mkb_percentage": wv.get("mkb_percentage", 0.14),
            "belastbare_winst": wv.get("belastbare_winst", 0),
            "omzet_per_klant": wv.get("omzet_per_klant", []),
            "kosten_per_categorie": wv.get("kosten_per_categorie", []),
        },
        "balans": {
            "activa": {
                "mva": bal.get("activa_mva", {"begin": 0, "eind": 0}),
                "debiteuren": bal.get("activa_debiteuren", {"begin": 0, "eind": 0}),
                "liquide_middelen": bal.get("activa_liquide_middelen", {"begin": None, "eind": None}),
                "totaal": bal.get("activa_totaal", {"begin": 0, "eind": 0}),
            },
            "passiva": {
                "eigen_vermogen": bal.get("passiva_eigen_vermogen", {"begin": 0, "eind": 0}),
                "crediteuren": bal.get("passiva_crediteuren", {"begin": 0, "eind": 0}),
                "btw_schuld": bal.get("passiva_btw_schuld", {"begin": 0, "eind": 0}),
                "kortlopend_totaal": bal.get("passiva_kortlopend", {"begin": 0, "eind": 0}),
                "totaal": bal.get("passiva_totaal", {"begin": 0, "eind": 0}),
            },
        },
        "mva": {
            "items": mva_data.get("items", []),
            "totaal_boekwaarde_begin": mva_data.get("totaal_boekwaarde_begin", 0),
            "totaal_boekwaarde_eind": mva_data.get("totaal_boekwaarde_eind", 0),
            "totaal_afschrijving": mva_data.get("totaal_afschrijving", 0),
            "totaal_aanschaf_dit_jaar": mva_data.get("totaal_aanschaf_dit_jaar", 0),
        },
        "bron": "accountant",
    }


def _get_prev_year_eind(overrides: dict[int, dict], jaar: int) -> dict | None:
    """Get the eind values from the previous year's override for use as begin values."""
    prev = overrides.get(jaar - 1)
    if not prev:
        return None

    bal = prev.get("balans", {})
    return {
        "mva": bal.get("activa_mva", {}).get("eind", 0),
        "debiteuren": bal.get("activa_debiteuren", {}).get("eind", 0),
        "liquide_middelen": bal.get("activa_liquide_middelen", {}).get("eind"),
        "crediteuren": bal.get("passiva_crediteuren", {}).get("eind", 0),
        "btw_schuld": bal.get("passiva_btw_schuld", {}).get("eind", 0),
    }


# === Endpoints ===

@router.get("/bank-status")
async def get_bank_status(user: dict = Depends(get_current_user)):
    """Get overview of uploaded bank CSVs."""
    db = get_db()
    uid = user["uid"]
    docs = list(
        db.collection("bank_accounts")
        .where(filter=FieldFilter("user_id", "==", uid))
        .stream()
    )
    accounts = []
    for doc in docs:
        d = doc.to_dict()
        accounts.append({
            "id": doc.id,
            "account_number": d.get("account_number", ""),
            "account_name": d.get("account_name", ""),
            "min_date": d.get("min_date", ""),
            "max_date": d.get("max_date", ""),
            "transaction_count": d.get("transaction_count", 0),
            "uploaded_at": d.get("uploaded_at", ""),
        })
    return {"accounts": accounts}


@router.post("/upload-csv")
async def upload_bank_csv(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """Upload and process an ING bank CSV (betaalrekening or spaarrekening)."""
    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(400, "Alleen CSV-bestanden zijn toegestaan")

    content = (await file.read()).decode("utf-8-sig")
    try:
        account_name, account_number, transactions = _detect_and_parse_csv(content)
    except Exception as e:
        raise HTTPException(400, f"Fout bij verwerken CSV: {str(e)}")

    if not transactions:
        raise HTTPException(400, "Geen transacties gevonden in CSV")

    db = get_db()
    uid = user["uid"]

    # Find date range
    dates = [tx["datum"] for tx in transactions if tx["datum"]]
    min_date = min(dates) if dates else ""
    max_date = max(dates) if dates else ""

    # Delete existing transactions for this account
    existing = list(
        db.collection("bank_transactions")
        .where(filter=FieldFilter("user_id", "==", uid))
        .where(filter=FieldFilter("account_number", "==", account_number))
        .stream()
    )
    batch = db.batch()
    for doc in existing:
        batch.delete(doc.reference)
    batch.commit()

    # Store new transactions in batches of 500
    for i in range(0, len(transactions), 500):
        batch = db.batch()
        for tx in transactions[i:i + 500]:
            ref = db.collection("bank_transactions").document()
            batch.set(ref, {
                **tx,
                "account_number": account_number,
                "user_id": uid,
            })
        batch.commit()

    # Upsert account metadata
    existing_acc = list(
        db.collection("bank_accounts")
        .where(filter=FieldFilter("user_id", "==", uid))
        .where(filter=FieldFilter("account_number", "==", account_number))
        .stream()
    )
    acc_data = {
        "account_number": account_number,
        "account_name": account_name,
        "min_date": min_date,
        "max_date": max_date,
        "transaction_count": len(transactions),
        "uploaded_at": date.today().isoformat(),
        "user_id": uid,
    }
    if existing_acc:
        existing_acc[0].reference.update(acc_data)
    else:
        db.collection("bank_accounts").add(acc_data)

    return {
        "account_name": account_name,
        "account_number": account_number,
        "transactions": len(transactions),
        "min_date": min_date,
        "max_date": max_date,
    }


@router.delete("/bank/{account_id}")
async def delete_bank_account(
    account_id: str,
    user: dict = Depends(get_current_user),
):
    """Delete a bank account and its transactions."""
    db = get_db()
    uid = user["uid"]

    doc = db.collection("bank_accounts").document(account_id).get()
    if not doc.exists or doc.to_dict().get("user_id") != uid:
        raise HTTPException(404, "Bankrekening niet gevonden")

    account_number = doc.to_dict().get("account_number", "")

    # Delete transactions
    txs = list(
        db.collection("bank_transactions")
        .where(filter=FieldFilter("user_id", "==", uid))
        .where(filter=FieldFilter("account_number", "==", account_number))
        .stream()
    )
    batch = db.batch()
    for tx_doc in txs:
        batch.delete(tx_doc.reference)
    batch.commit()

    # Delete account
    db.collection("bank_accounts").document(account_id).delete()
    return {"ok": True}


@router.get("/overzicht")
async def get_jaarcijfers_overzicht(
    user: dict = Depends(get_current_user),
):
    """Get jaarcijfers for ALL available years in one call."""
    db = get_db()
    uid = user["uid"]

    all_invoice_data, all_expense_data, beschikbare_jaren = _load_all_data(db, uid)
    bank_accounts = _load_bank_data(db, uid)
    overrides = _load_overrides(db, uid)

    # Add override years to beschikbare_jaren
    for y in overrides:
        if y not in beschikbare_jaren:
            beschikbare_jaren.append(y)
    beschikbare_jaren = sorted(set(beschikbare_jaren), reverse=True)

    # Compute for each year
    jaren_data = {}
    for y in sorted(beschikbare_jaren):
        if y in overrides:
            # Use accountant data
            jaren_data[y] = _override_to_jaarcijfers(overrides[y])
        else:
            # Compute, using previous year's override eind as begin
            prev_eind = _get_prev_year_eind(overrides, y)
            jaren_data[y] = _compute_jaarcijfers(
                y,
                all_invoice_data,
                all_expense_data,
                bank_accounts,
                prev_eind,
            )

    # Bank status summary
    bank_status = []
    for acc_nr, acc in bank_accounts.items():
        bank_status.append({
            "account_number": acc_nr,
            "account_name": acc.get("name", acc_nr),
            "min_date": acc.get("min_date", ""),
            "max_date": acc.get("max_date", ""),
        })

    return {
        "beschikbare_jaren": beschikbare_jaren,
        "jaren": jaren_data,
        "bank_status": bank_status,
    }


@router.get("/{jaar}")
async def get_jaarcijfers(
    jaar: int,
    user: dict = Depends(get_current_user),
):
    """Generate full jaarcijfers for a given year."""
    db = get_db()
    uid = user["uid"]

    all_invoice_data, all_expense_data, beschikbare_jaren = _load_all_data(db, uid)
    overrides = _load_overrides(db, uid)

    # Add override years
    for y in overrides:
        if y not in beschikbare_jaren:
            beschikbare_jaren.append(y)
    beschikbare_jaren = sorted(set(beschikbare_jaren), reverse=True)

    if jaar in overrides:
        result = _override_to_jaarcijfers(overrides[jaar])
    else:
        bank_accounts = _load_bank_data(db, uid)
        prev_eind = _get_prev_year_eind(overrides, jaar)
        result = _compute_jaarcijfers(jaar, all_invoice_data, all_expense_data, bank_accounts, prev_eind)

    result["beschikbare_jaren"] = beschikbare_jaren
    return result


def _filename_from_url(url: str) -> str:
    """Extract the original filename from a Firebase Storage public URL."""
    path = unquote(urlparse(url).path)
    # Storage URLs: /v0/b/bucket/o/expenses%2Fuid%2Ftimestamp_filename.pdf
    # After unquote the path contains the full storage path
    return path.rsplit("/", 1)[-1] if "/" in path else path


def _date_to_dutch(date_str: str) -> str:
    """Convert YYYY-MM-DD to DD-MM-YYYY."""
    try:
        parts = date_str.split("-")
        if len(parts) == 3 and len(parts[0]) == 4:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
    except Exception:
        pass
    return date_str


def _storage_path_from_url(url: str) -> str:
    """Extract the Firebase Storage blob path from a public URL.
    
    Public URLs look like:
      https://storage.googleapis.com/BUCKET/expenses/uid/timestamp_file.pdf
    Or Firebase REST API URLs:
      https://firebasestorage.googleapis.com/v0/b/BUCKET/o/path%2Fto%2Ffile
    """
    if not url:
        return ""
    decoded = unquote(urlparse(url).path)
    # REST API format: /v0/b/bucket/o/the/actual/path
    if "/o/" in decoded:
        return decoded.split("/o/", 1)[1]
    # Public URL format: /BUCKET/the/actual/path
    bucket = FIREBASE_STORAGE_BUCKET
    if bucket and f"/{bucket}/" in decoded:
        return decoded.split(f"/{bucket}/", 1)[1]
    # Fallback: everything after 3rd slash segment
    parts = decoded.lstrip("/").split("/", 1)
    return parts[1] if len(parts) > 1 else ""


@router.get("/{jaar}/export")
async def export_jaarcijfers(
    jaar: int,
    user: dict = Depends(get_current_user),
):
    """Export a ZIP with Excel overview + all invoice/expense PDFs for the year."""
    db = get_db()
    uid = user["uid"]

    all_invoice_data, all_expense_data, _ = _load_all_data(db, uid)

    # Load company settings + customers for on-the-fly PDF generation
    settings_doc = db.collection("company_settings").document(uid).get()
    company = settings_doc.to_dict() if settings_doc.exists else {}

    # Pre-load all customers
    customer_cache = {}
    for cdoc in db.collection("customers").where(filter=FieldFilter("user_id", "==", uid)).stream():
        customer_cache[cdoc.id] = cdoc.to_dict()

    # Filter invoices for this year (status verzonden/betaald)
    year_invoices = [
        inv for inv in all_invoice_data
        if get_year(inv.get("factuurdatum", "")) == jaar
        and inv.get("status") in ("verzonden", "betaald")
    ]

    # Filter expenses for this year
    year_expenses = [
        exp for exp in all_expense_data
        if get_year(exp.get("datum", "")) == jaar
    ]

    # Build rows for Excel
    rows = []
    bucket = storage.bucket()

    for inv in year_invoices:
        factuurnummer = inv.get("factuurnummer", "")
        bestand = f"{factuurnummer}.pdf"

        # Try storage path from pdf_url first, fallback to convention
        storage_path = _storage_path_from_url(inv.get("pdf_url", ""))
        if not storage_path:
            storage_path = f"invoices/{uid}/{inv.get('id')}.pdf"

        # Try to get existing PDF; generate on-the-fly if missing
        pdf_bytes = None
        try:
            blob = bucket.blob(storage_path)
            pdf_bytes = blob.download_as_bytes()
        except Exception:
            pass

        if not pdf_bytes:
            # Generate PDF on the fly
            try:
                klant = customer_cache.get(inv.get("klant_id", ""), {})
                pdf_bytes = generate_invoice_pdf(inv, company, klant)
            except Exception:
                pdf_bytes = None
        rows.append({
            "in_uit": "In",
            "factuurdatum": inv.get("factuurdatum", ""),
            "factuurnummer": factuurnummer,
            "daan_of_wim": inv.get("daan_of_wim", "") or "",
            "btw": inv.get("btw_totaal", 0),
            "waarde": inv.get("totaal", 0),
            "bestand": bestand,
            "pdf_bytes": pdf_bytes,
        })

    for exp in year_expenses:
        pdf_url = exp.get("pdf_url", "") or ""
        # Determine filename for the PDF
        if pdf_url:
            raw_name = _filename_from_url(pdf_url)
            # Strip the timestamp prefix (2024-01-01T00:00:00+00:00_filename.pdf)
            if "_" in raw_name:
                bestand = raw_name.split("_", 1)[1]
            else:
                bestand = raw_name
        else:
            bestand = ""

        storage_path = _storage_path_from_url(pdf_url)

        # Download expense PDF
        pdf_bytes = None
        if storage_path:
            try:
                blob = bucket.blob(storage_path)
                pdf_bytes = blob.download_as_bytes()
            except Exception:
                pass

        rows.append({
            "in_uit": "Uit",
            "factuurdatum": exp.get("datum", ""),
            "factuurnummer": exp.get("factuurnummer", ""),
            "daan_of_wim": exp.get("daan_of_wim", "") or "",
            "btw": -(exp.get("btw", 0) or 0),
            "waarde": -(exp.get("totaal", 0) or 0),
            "bestand": bestand,
            "pdf_bytes": pdf_bytes,
        })

    # Sort by date descending
    rows.sort(key=lambda r: r["factuurdatum"], reverse=True)

    # === Build Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = str(jaar)

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Overzicht facturen {jaar} – Opwolken.com VOF"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="left")

    # Header row (row 3)
    headers = ["In/Uit", "Factuurdatum", "Factuurnummer", "Daan of Wim", "BTW", "Waarde", "Bestand"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)

    # Data rows
    for i, row in enumerate(rows, 4):
        ws.cell(row=i, column=1, value=row["in_uit"])
        ws.cell(row=i, column=2, value=_date_to_dutch(row["factuurdatum"]))
        ws.cell(row=i, column=3, value=row["factuurnummer"])
        ws.cell(row=i, column=4, value=row["daan_of_wim"])
        btw_cell = ws.cell(row=i, column=5, value=row["btw"])
        btw_cell.number_format = '#,##0.00'
        val_cell = ws.cell(row=i, column=6, value=row["waarde"])
        val_cell.number_format = '#,##0.00'
        ws.cell(row=i, column=7, value=row["bestand"])

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, "column_letter"):
                col_letter = cell.column_letter
            if cell.value and not isinstance(cell, type(None)):
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    # === Build ZIP ===
    zip_buf = io.BytesIO()

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"overzicht-{jaar}.xlsx", excel_buf.read())

        # Add PDFs
        for row in rows:
            if not row["bestand"] or not row.get("pdf_bytes"):
                continue
            folder = "inkomsten" if row["in_uit"] == "In" else "uitgaven"
            zf.writestr(f"{folder}/{row['bestand']}", row["pdf_bytes"])

    zip_buf.seek(0)

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="boekhouding-{jaar}.zip"'
        },
    )
